"""Excel export utilities for disclosure scoring outputs."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, Sequence, cast

import pandas as pd

from config.settings import DISCLOSURE_SCORES_FILE, EXCEL_EXPORT_CONFIG
from src.utils.logger import get_logger


logger = get_logger(__name__)


@dataclass(frozen=True)
class ExcelExporterConfig:
    """Configuration for Excel exports."""

    output_path: Path = DISCLOSURE_SCORES_FILE
    engine: str = str(EXCEL_EXPORT_CONFIG.get("engine", "openpyxl"))
    index: bool = bool(EXCEL_EXPORT_CONFIG.get("index", False))
    freeze_panes: tuple[int, int] = cast(
        tuple[int, int],
        EXCEL_EXPORT_CONFIG.get("freeze_panes", (1, 0)),
    )
    auto_filter: bool = bool(EXCEL_EXPORT_CONFIG.get("auto_filter", True))


class ExcelExporter:
    """Export disclosure scores, category counts, and validation sheets."""

    SHEET_DISCLOSURE_SCORES = "Disclosure Scores"
    SHEET_CATEGORY_COUNTS = "Category Counts"
    SHEET_SECTION_SUMMARY = "Section Summary"
    SHEET_VALIDATION_READY = "Validation Ready"

    def __init__(self, config: ExcelExporterConfig | None = None) -> None:
        """Initialize an Excel exporter."""
        self.config = config or ExcelExporterConfig()

    def export_scores(
        self,
        scores: Mapping[str, object] | Sequence[Mapping[str, object]],
        output_path: str | Path | None = None,
    ) -> Path:
        """Export disclosure scores to a workbook."""
        return self.export_complete_workbook(
            scores=scores,
            categories=None,
            sections=None,
            validation=None,
            output_path=output_path,
        )

    def export_categories(
        self,
        categories: Mapping[str, object] | Sequence[Mapping[str, object]],
        output_path: str | Path | None = None,
    ) -> Path:
        """Export category counts to a workbook."""
        return self.export_complete_workbook(
            scores=None,
            categories=categories,
            sections=None,
            validation=None,
            output_path=output_path,
        )

    def export_sections(
        self,
        sections: Mapping[str, str] | Sequence[Mapping[str, object]],
        output_path: str | Path | None = None,
    ) -> Path:
        """Export section summaries to a workbook."""
        return self.export_complete_workbook(
            scores=None,
            categories=None,
            sections=sections,
            validation=None,
            output_path=output_path,
        )

    def export_validation(
        self,
        validation: Mapping[str, object] | Sequence[Mapping[str, object]],
        output_path: str | Path | None = None,
    ) -> Path:
        """Export validation-ready records to a workbook."""
        return self.export_complete_workbook(
            scores=None,
            categories=None,
            sections=None,
            validation=validation,
            output_path=output_path,
        )

    def export_complete_workbook(
        self,
        scores: Mapping[str, object] | Sequence[Mapping[str, object]] | None,
        categories: Mapping[str, object] | Sequence[Mapping[str, object]] | None,
        sections: Mapping[str, str] | Sequence[Mapping[str, object]] | None,
        validation: Mapping[str, object] | Sequence[Mapping[str, object]] | None,
        output_path: str | Path | None = None,
    ) -> Path:
        """Export a complete multi-sheet workbook.

        Args:
            scores: Disclosure score result or score records.
            categories: Category statistics mapping or flat records.
            sections: Extracted section text mapping or section records.
            validation: Validation-ready mapping or flat records.
            output_path: Optional destination workbook path.

        Returns:
            Path to the exported workbook.
        """
        destination = Path(output_path) if output_path else self.config.output_path
        if destination.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError(
                "Excel output path must use an .xlsx or .xlsm extension.",
            )
        destination.parent.mkdir(parents=True, exist_ok=True)

        try:
            frames = {
                self.SHEET_DISCLOSURE_SCORES: self._scores_to_frame(scores),
                self.SHEET_CATEGORY_COUNTS: self._categories_to_frame(categories),
                self.SHEET_SECTION_SUMMARY: self._sections_to_frame(sections),
                self.SHEET_VALIDATION_READY: self._validation_to_frame(
                    validation,
                    scores,
                    categories,
                    sections,
                ),
            }

            with pd.ExcelWriter(destination, engine=self.config.engine) as writer:
                for sheet_name, frame in frames.items():
                    frame.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=self.config.index,
                    )

            self._format_workbook(destination, frames)
            logger.info("Workbook created: %s", destination)
            logger.info("Excel exported: %s", destination)
            return destination
        except ModuleNotFoundError as exc:
            logger.exception("Excel export requires openpyxl: %s", exc)
            raise
        except Exception as exc:
            logger.exception("Errors encountered during Excel export: %s", exc)
            raise

    def _scores_to_frame(
        self,
        scores: Mapping[str, object] | Sequence[Mapping[str, object]] | None,
    ) -> pd.DataFrame:
        """Convert disclosure scores into an export dataframe."""
        if scores is None:
            return pd.DataFrame()
        if isinstance(scores, Sequence) and not isinstance(scores, Mapping):
            return pd.DataFrame(list(scores))

        score_mapping = dict(scores)
        component_scores = score_mapping.pop("component_scores", {})
        raw_metrics = score_mapping.pop("raw_metrics", {})
        record = self._flatten_mapping(score_mapping)
        if isinstance(component_scores, Mapping):
            record.update(
                {
                    f"component_{key}": value
                    for key, value in component_scores.items()
                },
            )
        if isinstance(raw_metrics, Mapping):
            for key in ("total_keyword_count", "total_word_count", "keyword_density"):
                if key in raw_metrics:
                    record[key] = raw_metrics[key]
        return pd.DataFrame([record])

    def _categories_to_frame(
        self,
        categories: Mapping[str, object] | Sequence[Mapping[str, object]] | None,
    ) -> pd.DataFrame:
        """Convert category statistics into an export dataframe."""
        if categories is None:
            return pd.DataFrame()
        if isinstance(categories, Sequence) and not isinstance(categories, Mapping):
            return pd.DataFrame(list(categories))

        records: list[dict[str, object]] = []
        for category, values in categories.items():
            if isinstance(values, Mapping):
                record = {"category": category}
                record.update(values)
            else:
                record = {"category": category, "count": values}
            records.append(record)
        return pd.DataFrame(records)

    def _sections_to_frame(
        self,
        sections: Mapping[str, str] | Sequence[Mapping[str, object]] | None,
    ) -> pd.DataFrame:
        """Convert section text into compact section summary records."""
        if sections is None:
            return pd.DataFrame()
        if isinstance(sections, Sequence) and not isinstance(sections, Mapping):
            return pd.DataFrame(list(sections))

        records = []
        for section_name, section_text in sections.items():
            text = str(section_text or "")
            records.append(
                {
                    "section": section_name,
                    "character_count": len(text),
                    "word_count": len(text.split()),
                    "present": bool(text.strip()),
                    "preview": text[:500],
                },
            )
        return pd.DataFrame(records)

    def _validation_to_frame(
        self,
        validation: Mapping[str, object] | Sequence[Mapping[str, object]] | None,
        scores: Mapping[str, object] | Sequence[Mapping[str, object]] | None,
        categories: Mapping[str, object] | Sequence[Mapping[str, object]] | None,
        sections: Mapping[str, str] | Sequence[Mapping[str, object]] | None,
    ) -> pd.DataFrame:
        """Build validation-ready export records."""
        if validation is not None:
            if isinstance(validation, Sequence) and not isinstance(validation, Mapping):
                return pd.DataFrame(list(validation))
            return pd.DataFrame([self._flatten_mapping(validation)])

        score_frame = self._scores_to_frame(scores)
        category_frame = self._categories_to_frame(categories)
        section_frame = self._sections_to_frame(sections)
        return pd.DataFrame(
            [
                {
                    "score_records": len(score_frame),
                    "category_records": len(category_frame),
                    "section_records": len(section_frame),
                    "ready_for_validation": not score_frame.empty,
                },
            ],
        )

    def _format_workbook(
        self,
        workbook_path: Path,
        frames: Mapping[str, pd.DataFrame],
    ) -> None:
        """Apply light openpyxl formatting to exported workbook."""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        workbook = load_workbook(workbook_path)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        for sheet_name, frame in frames.items():
            worksheet = workbook[sheet_name]
            freeze_row, freeze_column = self.config.freeze_panes
            if freeze_row > 0 or freeze_column > 0:
                worksheet.freeze_panes = worksheet.cell(
                    row=max(freeze_row + 1, 1),
                    column=max(freeze_column + 1, 1),
                )
            else:
                worksheet.freeze_panes = None

            if not frame.empty and self.config.auto_filter:
                worksheet.auto_filter.ref = worksheet.dimensions

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            for column_cells in worksheet.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                    max(max_length + 2, 12),
                    60,
                )

        workbook.save(workbook_path)

    def _flatten_mapping(
        self,
        mapping: Mapping[str, object],
        prefix: str = "",
    ) -> dict[str, object]:
        """Flatten nested mappings for tabular export."""
        flattened: dict[str, object] = {}
        for key, value in mapping.items():
            full_key = f"{prefix}{key}" if not prefix else f"{prefix}_{key}"
            if isinstance(value, Mapping):
                flattened.update(self._flatten_mapping(value, full_key))
            elif isinstance(value, (list, tuple, set)):
                flattened[full_key] = ", ".join(str(item) for item in value)
            else:
                flattened[full_key] = value
        return flattened
